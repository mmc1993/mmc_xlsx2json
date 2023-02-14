# coding=utf-8

#   作者: mmc
#   日期: 2020/3/19

from signal import valid_signals
import sys
import time
import openpyxl

def is_float(str):
    try:
        float(str)
    except ValueError:
        return False
    return True

#   读取单元格文本
def XLGetValue(xlsx, cur_row, cur_col):
    return str(xlsx.cell(cur_row, cur_col).value)

def ValToKey(value):
    return value[0] != "\"" and "\"" + value + "\"" or value

#   跳过空字符
def SkipSpaceChar(buffer, i, l):
    while i != l and ord(buffer[i]) <= 32:
        i = i + 1
    return i

#   跳过注释行
def SkipCommentLine(xlsx, cur_row):
    while cur_row != xlsx.max_row and XLGetValue(xlsx, cur_row, 1) == "//":
        cur_row = cur_row + 1
    return cur_row

#   type name => (type, name)
def SplitTypeName(value):
    split = value.split()
    return len(split) == 1      \
        and (split[0], None)    \
        or  (split[0], split[1])

#   字符串比较
def MatchSubString(str0, str1, i):
    l0 = len(str0)
    l1 = len(str1)
    if l0 >= l1:
        sub = str0[i:i+l1]
        return sub == str1
    return False

#   解析器
class Parser:
    def __init__(self, func = None, name = None):
        self.mFunc = func
        self.mName = name
        self.mChildren = []
    
    def Append(self, unit):
        self.mChildren.append(unit)

    def SetName(self, name):
        self.mName = name

    def GetName(self):
        return self.mName
    
    def GetType(self):
        if   self.mFunc == OnParseList:
            return "list"
        elif self.mFunc == OnParseDict:
            return "dict"
        elif self.mFunc == OnParseBool:
            return "bool"
        elif self.mFunc == OnParseInt:
            return "int"
        elif self.mFunc == OnParseStr:
            return "str"
        elif self.mFunc == OnParseFloat:
            return "float"
        elif self.mFunc == OnParseStruct:
            return "struct"

    def GetChildren(self):
        return self.mChildren

    def Run(self, value, i, l):
        return self.mFunc(value, i, l, self.mChildren)

def OnParseBool(value, i, l, parser):
    assert is_float(value[i]), value
    return i + 1, value[i] == "0"       \
                          and "false"   \
                          or  "true"

def OnParseInt(value, i, l, parser):
    num = []
    while i != l:
        if "0" <= value[i] and "9" >= value[i]:
            num.append(value[i]); i = i + 1
        elif len(num) == 0 and "-" == value[i]:
            num.append(value[i]); i = i + 1
        else:
            break
    return i, "".join(num)

def OnParseStr(value, i, l, parser):
    assert value[i] == "\"", value
    i = i + 1
    str = [ ]
    mis = False
    while i != l:
        if value[i] == "\\": mis = True
        if value[i] == "\"" and not mis:
            break
        str.append(value[i])
        i   = i + 1
        mis = False
    assert value[i] == "\"", value
    return i + 1, "\"" + "".join(str) + "\""

def OnParseFloat(value, i, l, parser):
    ret = []
    dot = False
    while i != l:
        if "0" <= value[i] and "9" >= value[i]:
            ret.append(value[i]); i = i + 1
        elif len(ret) == 0 and "-" == value[i]:
            ret.append(value[i]); i = i + 1
        elif not dot and "." == value[i]:
            ret.append(value[i]); i = i + 1
        else:
            break
    return i, "".join(ret)

def OnParseList(value, i, l, parser):
    assert value[i] == "[", value
    ret = []
    while i != l:
        if value[i] == "]": break
        i       = SkipSpaceChar(value, i + 1, l)
        if value[i] == "]": break

        i, data = parser[0].Run(value, i, l)
        i       = SkipSpaceChar(value, i, l)
        ret.append(data)
        assert value[i] == "," \
            or value[i] == "]", value
    assert value[i] == "]", value
    return i + 1, "[" + ", ".join(ret) + "]"

def OnParseDict(value, i, l, parser):
    assert value[i] == "{", value
    ret = []
    while i != l:
        if value[i] == "}": break
        i = SkipSpaceChar(value, i + 1, l)
        if value[i] == "}": break

        i, key = parser[0].Run(value, i, l)
        i      = SkipSpaceChar(value, i, l)
        assert value[i] == ":", value
        i      = SkipSpaceChar(value, i + 1, l)
        i, val = parser[1].Run(value, i, l)
        ret.append("%s: %s" %(key, val))
        assert value[i] == "," \
            or value[i] == "}", value
    assert value[i] == "}", value
    return i + 1, "{" + ", ".join(ret) + "}"

def OnParseStruct(value, i, l, parser):
    assert value[i] == "<", value
    ret = []
    idx = 0
    while i != l:
        if value[i] == ">": break
        i = SkipSpaceChar(value, i + 1, l)
        key    = parser[idx].GetName()
        i, val = parser[idx].Run(value, i, l)
        i = SkipSpaceChar(value, i, l)
        ret.append("\"%s\": %s" % (key, val))
        idx = idx + 1
        assert value[i] == "," \
            or value[i] == ">", value
    assert value[i] == ">", value
    return i + 1, "{" + ", ".join(ret) + "}"

def GetTypeName(value, i, l):
    b = i
    while i != l and (\
        ("a" <= value[i] and "z" >= value[i]) or \
        ("A" <= value[i] and "Z" >= value[i]) or \
        ("0" <= value[i] and "9" >= value[i]) or \
        ("_" == value[i])):
        i = i + 1
    return i, value[b: i]

def CreateParser(value, i, l):
    parser = None
    if   MatchSubString(value, "bool", i):
        i = i + len("bool")
        parser = Parser(OnParseBool)

    elif MatchSubString(value, "int", i):
        i = i + len("int")
        parser = Parser(OnParseInt)

    elif MatchSubString(value, "str", i):
        i = i + len("str")
        parser = Parser(OnParseStr)

    elif MatchSubString(value, "float", i):
        i = i + len("float")
        parser = Parser(OnParseFloat)

    elif value[i] == "[":
        i = i + 1
        i = SkipSpaceChar(value, i, l)
        i, valParser = CreateParser(value, i, l)
        i = SkipSpaceChar(value, i, l)
        parser = Parser(OnParseList)
        parser.Append(valParser)
        assert value[i] == "]", value
        i = i + 1

    elif value[i] == "{":
        i = i + 1
        keyParser = Parser(OnParseStr)
        i = SkipSpaceChar(value, i, l)
        i, valParser = CreateParser(value, i, l)
        i = SkipSpaceChar(value, i, l)
        parser = Parser(OnParseDict)
        parser.Append(keyParser)
        parser.Append(valParser)
        assert value[i] == "}", value
        i = i + 1

    elif value[i] == "<":
        parser = Parser(OnParseStruct)
        while i != l:
            if value[i] == ">": break
            i = i + 1
            i = SkipSpaceChar(value, i, l)
            i, valParser = CreateParser(value, i, l)
            i = SkipSpaceChar(value, i, l)
            parser.Append(valParser)

            assert value[i] == "," \
                or value[i] == ">", value
        assert value[i] == ">", value
        i = i + 1

    i = SkipSpaceChar(value, i, l)
    i,k = GetTypeName(value, i, l)
    parser.SetName(k)
    return i, parser

#   初始化解析器
def CreateParserList(xlsx, cur_row, parser_list):
    for cur_col in range(1, xlsx.max_column + 1):
        try:
            cur_val = XLGetValue(xlsx, cur_row, cur_col)
            _,v = CreateParser(cur_val, 0, len(cur_val))
            parser_list.append(v)
        except AssertionError as e:
            assert False, "%d:%d | %s" % (cur_row, cur_col, e)
    return cur_row + 1

#   解析单元格
def CreateOutputList(xlsx, cur_row, parser_list, out):
    for cur_row in range(cur_row, xlsx.max_row + 1):
        if XLGetValue(xlsx, cur_row, 1) == "//": continue

        lines = []
        for cur_col in range(1, xlsx.max_column + 1):
            try:
                value = XLGetValue(xlsx, cur_row, cur_col)
                vlen = len(value)
                parser = parser_list[cur_col - 1]
                k    = parser.GetName()
                _, v = parser.Run(value, 0, vlen)
                lines.append("\"%s\": %s" % (k, v))
            except AssertionError as e:
                assert False, "%d:%d | %s" % (cur_row, cur_col, e)
        key = XLGetValue(xlsx, cur_row, 1)
        val = "{" + ", ".join(lines) + "}"
        out.append("%s: %s" % (ValToKey(key), val))
    return cur_row

#   导出Json
def ToJson(ifile):
    # clock = time.time()
    #   读第一张表
    xlsx = openpyxl.load_workbook(ifile, data_only = True)
    xlsx = xlsx[xlsx.sheetnames[0]]

    #   跳过注释行
    parser_list = []
    output_list = []
    cur_row = SkipCommentLine(xlsx, 1)

    try:
        print("> Export %s" % ifile)
        cur_row = CreateParserList(xlsx, cur_row, parser_list)
        cur_row = CreateOutputList(xlsx, cur_row, parser_list, output_list)
        # print("> %.3fs From %s" % (time.time() - clock, ifile))
        return "{\n" + ",\n".join(output_list) + "\n}", parser_list
    except AssertionError as e:
        assert False, "%s | %s" % (ifile, e)

def ExportToFile(ifile, ofile):
    with open(ofile, "w") as f:
        f.write((ToJson(ifile))[0])

if __name__ == "__main__":
    ExportToFile(sys.argv[1], sys.argv[2])
