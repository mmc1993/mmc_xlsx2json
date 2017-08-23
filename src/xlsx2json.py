# -*- coding:utf-8 -*- 

import openpyxl

def RetKey(key):
	if key[0] != "\""  or key[-1] != "\"":
		return "\"" + key + "\""
	return key

def RetArray(data):
	return "[" + data + "]"

def RetDict(data):
	return "{" + data + "}"

def ParseInt(data, parse):
	return data

def ParseBool(data, parse):
	return data == 0 and "false" or "true"

def ParseFloat(data, parse):
	if -1 == data.find("."):
		data = data + ".0"
	return data

def ParseString(data, parse):
	return RetKey(data.replace("\n", "\\n").replace("\"", "\\\""))

def ParseStruct(data, parse):
	result = []
	for i, v in enumerate(data.split(";")[:-1]):
		key = parse[i]["name"]
		val = parse[i]["func"](v, parse[i]["childs"])
		result.append("%s: %s" % (RetKey(key), val))
	return RetDict(", ".join(result))

def ParseList(data, parse):
	result = []
	for k, v in enumerate(data.split(",")[:-1]):
		result.append(parse[0]["func"](v, parse[0]["childs"]))
	return RetArray(",".join(result))

def GetValue(xl, row, col):
	return unicode(xl.cell(row = row, column = col).value or "")

def GetLine(xl, row):
	return [GetValue(xl, row, col) for col in xrange(1, xl.max_column + 1)]

def GetRecord(parses, line):
	chunk = []
	for i, raw in enumerate(line):
		if len(raw) != 0:
			parse = parses[i]
			key = parse["name"]
			val = parse["func"](raw, parse["childs"])
			chunk.append("%s: %s" % (RetKey(key), val))
	return RetDict(", ".join(chunk))

def CheckParse(type, name):
	result = {"name": None, "func": None, "childs": None}
	if type[0:len("int")] == "int":
		result["func"] = ParseInt
	elif type[0:len("bool")] == "bool":
		result["func"] = ParseBool
	elif type[0:len("float")] == "float":
		result["func"] = ParseFloat
	elif type[0:len("string")] == "string":
		result["func"] = ParseString
	elif type[0:len("struct")] == "struct":
		result["func"] = ParseStruct
		result["childs"] = CheckParses(type[len("struct")+1:].split(","))
	elif type[0:len("list")] == "list":
		result["func"] = ParseList
		result["childs"] = CheckParses([type[len("list")+1:]])
	if len(name) > 0:
		result["name"] = RetKey(name)
	return result

def CheckParses(descs):
	result = []
	for desc in descs:
		pos = desc.rfind(" ")
		type, name = desc, ""
		if -1 != pos:
			type = desc[0: pos]
			name = desc[pos+1:]
		result.append(CheckParse(type, name))
	return result

def CheckJson(xl):
	result = []
	#	第一行注释
	#	第二行类型
	parses = CheckParses(GetLine(xl, 2))
	#	第三行内容
	for row in xrange(3, xl.max_row + 1):
		key = parses[0]["func"](GetValue(xl, row, 1), parses[0]["childs"])
		val = GetRecord(parses, GetLine(xl, row))
		result.append("%s: %s" % (RetKey(key), val))
	return RetDict(", \n".join(result))

def ToJson(xl):
	return CheckJson(xl.get_sheet_by_name(xl.sheetnames[0]))

if __name__ == "__main__":
	print ToJson(openpyxl.load_workbook("a.xlsx"))
