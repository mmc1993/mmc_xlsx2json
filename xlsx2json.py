# -*- coding:utf-8 -*- 

#   支持类型
#   int bool float string [] <>

import os
import sys
import time
import codecs
import openpyxl
import datetime

class Error(BaseException):
    def __init__(self, what):
        BaseException.__init__(self)
        self._what = what

    def What(self):
        return self._what

class Parser:
    def __init__(self, name = None, parse = None, *child):
        self.mName = name
        self.mChild = child
        self.mParse = parse

def Skip(buffer, i, l):
    while i != l and ord(buffer[i]) <= 32:
        i = i + 1
    return i

def IsSame(src, dst):
    return src[0: len(dst)] == dst

#   返回字符串
def RetStr(val):
    return '\"' + val + '\"'

#   返回key
def RetKey(key):
    if key[0] != '\"' or key[-1] != '\"':
        return RetStr(key)
    return key

#   返回字典
def RetDict(val):
    return '{' + val + '}'

#   返回数组
def RetList(val):
    return '[' + val + ']'

#   解析布尔
def ParseBool(buffer, i, l, parser):
    assert(buffer[i] == '1' or buffer[i] == '0')
    return (i + 1, buffer[i] == '1' and "true" or "false")

#   解析浮点数
def ParseFloat(buffer, i, l, parser):
    buf = []
    if (buffer[i] == '-'):
        buf.append(buffer[i])
        i = i + 1

    assert(i != l)
    while i != l \
        and (ord(buffer[i]) <= ord('9') \
        and ord(buffer[i]) >= ord('0') \
        or ord(buffer[i]) == ord('.')):
        buf.append(buffer[i])
        i = i + 1
    val = "".join(buf)
    float(val)
    return (i, val)

#   解析字符串
def ParseString(buffer, i, l, parser):
    assert(buffer[i] == '\"')
    buf = []
    while i != l:
        i = i + 1
        if buffer[i] == '\"':
            break
        if buffer[i] == '\n':
            buf.append("\\n")
        elif buffer[i] == '\\':
            buf.append(buffer[i + 0])
            buf.append(buffer[i + 1])
            i = i + 1
        else:
            buf.append(buffer[i])
    return (i + 1, RetStr("".join(buf)))

#   解析数组
def ParseList(buffer, i, l, parser):
    assert(buffer[i] == '[')
    if buffer[i + 1] == ']':
        return (i + 2, RetList(""))

    buf = []
    cparser = parser.mChild[0]
    while i != l:
        # i = i + 1
        i = Skip(buffer, i + 1, l)
        i, val = cparser.mParse(buffer, i, l, cparser)
        buf.append(val)
        if buffer[i] == ',':
            i = i + 1
        elif buffer[i] == ']':
            break
    return (i + 1, RetList(", ".join(buf)))

#   解析自定义类型
def ParseType(buffer, i, l, parser):
    assert(buffer[i] == '<')
    buf = []
    while i != l:
        for _, v in enumerate(parser.mChild):
            i, val = v.mParse(buffer, i + 1, l, v)
            buf.append("%s: %s" % (RetKey(v.mName), val))
            if buffer[i] == ',':
                i = i + 1
            elif buffer[i] == '>':
                break
        if buffer[i] == '>':
            break
    return (i + 1, RetDict(", ".join(buf)))

def ParseDict(buffer, i, l, parser):
    assert(buffer[i] == '{')
    if (buffer[i + 1] == '}'):
        return (i + 2, RetDict(""))
    i = i + 1
    buf = []
    keyParse = parser.mChild[0]
    valParse = parser.mChild[1]
    while i != l:
        i, key = keyParse.mParse(buffer, i + 0, l, keyParse)
        i, val = valParse.mParse(buffer, i + 2, l, valParse)
        buf.append("%s: %s" % (RetKey(key), val))
        if buffer[i] == '}':
            break
        elif buffer[i] == ',':
            i = Skip(buffer, i + 1, l)
    return (i + 1, RetDict(", ".join(buf)))

### 从xlsx读取
def GetVal(xlsx, row, col):
    return unicode(xlsx.cell(row = row, column = col).value)

def GetFieldName(buffer, i, l):
    pos = i
    while pos != l and \
        (ord(buffer[pos]) >= ord('a') and ord(buffer[pos]) <= ord('z') \
        or ord(buffer[pos]) >= ord('A') and ord(buffer[pos]) <= ord('Z') \
        or ord(buffer[pos]) >= ord('0') and ord(buffer[pos]) <= ord('9') \
        or buffer[pos] == '_'):
        pos = pos + 1
    return (pos, buffer[i: pos])

def MakeParser(buffer, i, l):
    stype = buffer[i:]
    parser = Parser()
    if IsSame(stype, "bool"):
        i = i + 4
        parser.mParse = ParseBool
    elif IsSame(stype, "float"):
        i = i + 5
        parser.mParse = ParseFloat
    elif IsSame(stype, "string"):
        i = i + 6
        parser.mParse = ParseString
    elif IsSame(stype, '{'):
        i, keyParser = MakeParser(buffer, i + 1, l)
        i, valParser = MakeParser(buffer, i + 2, l)
        parser.mChild = (keyParser, valParser)
        parser.mParse = ParseDict
        i = i + 1
    elif IsSame(stype, '<'):
        i = i + 1
        cparsers = []
        while i != l:
            i, par = MakeParser(buffer, i, l)
            cparsers.append(par)
            if buffer[i] == ',':
                # i = i + 2
                # 支持换行定义
                # <float a,
                #  float b>
                i = Skip(buffer, i + 1, l)
            elif buffer[i] == '>':
                i = i + 1
                break
        parser.mParse = ParseType
        parser.mChild = tuple(cparsers)

    if buffer[i] == ' ':
        i = i + 1
    else:
        assert(buffer[i] == '[' and buffer[i + 1] == ']')
        parser = Parser(None, ParseList, parser)
        i = i + 3

    i, parser.mName = GetFieldName(buffer, i, l)
    return (i, parser)

def MakeParsers(xlsx, row):
    parsers = []
    for col in xrange(1, xlsx.max_column + 1):
        try:
            raw = GetVal(xlsx, row, col)
            _, parser = MakeParser(raw, 0, len(raw))
            parsers.append(parser)
        except:
            raise Error(u"类型定义错误: 行 %s, 列 %s" % (row, col))
    return parsers

def GetRecord(parsers, xlsx, row):
    chunk = []
    for col in xrange(1, xlsx.max_column + 1):
        try:
            parser = parsers[col - 1]
            raw = GetVal(xlsx, row, col)
            _, val = parser.mParse(raw, 0, len(raw), parser)
            chunk.append("%s: %s" % (RetKey(parser.mName), val))
        except:
            raise Error(u"内容错误: 行 %s, 列 %s, 字段 %s, 内容 %s" % (row, col, parser.mName, raw))
    return RetDict(", ".join(chunk))

def ToJson(xlsx):
    #	第一行注释
    result = []
	#	第二行类型
    parsers = MakeParsers(xlsx, 2)
	#	第三行内容
    for row in xrange(3, xlsx.max_row + 1):
        par0 = parsers[0]
        raw = GetVal(xlsx, row, 1)
        _, key = par0.mParse(raw, 0, len(raw), par0)
        val = GetRecord(parsers, xlsx, row)
        result.append("%s: %s" % (RetKey(key), val))
    return RetDict(",\n".join(result))

def Export(fname):
    try:
        xlsx = openpyxl.load_workbook(fname, data_only = True)
        sheet = xlsx[xlsx.sheetnames[0]]
        return ToJson(sheet)
    except Error, error:
        print u"表 %s => %s" % (fname, error.What())
    except:
        print u"表 %s => 未知错误" % (fname)

def Write(output, fname):
    base = os.path.basename(fname)
    name = os.path.splitext(base)[0]
    newname = output + "\\" + name + ".json"
    with codecs.open(newname, 'w', "utf-8") as f:
        f.write(Export(fname))

INPUT_PATH = "C:\\MyWork\\Git\\xlsx2json\\xlsx\\a.xlsx"
OUTPUT_PATH = "C:\\MyWork\\Git\\xlsx2json\\export\\"

Write(OUTPUT_PATH, INPUT_PATH)

