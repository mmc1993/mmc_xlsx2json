# MIT License
# 
# Copyright (c) 2024 mmc
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

# coding=utf-8

import openpyxl

#   读取单元格文本
def xl_value(xl, row, col):
    return str(xl.cell(row, col).value)

def to_key(value):
    return "\"" + value + "\"" if value[0] != "\"" else value

#   前缀比较
def match_prefix(str0, str1, i):
    l0 = len(str0)
    l1 = len(str1)
    if l0 - i >= l1:
        sub = str0[i:i+l1]
        return sub == str1
    return False

#   是否注释
def is_comment_char(xl, row, col):
    return xl_value(xl, row, col) == "//"

#   跳过空字符
def skip_space_char(buffer, i, l):
    while i != l and ord(buffer[i]) <= 32: i = i + 1
    return i

#   跳过注释行
def skip_comment_line(xl, row):
    while row != xl.max_row and is_comment_char(xl, row, 1): row = row + 1
    return row

#   type name => (type, name)
def split_type_name(value):
    split = value.split()
    return (split[0], None) if len(split) == 1 else (split[0], split[1])

#   获取有效类型名
def get_type_name(buffer, index, length):
    start = index
    while index != length and (                             \
        ("a" <= buffer[index] and "z" >= buffer[index]) or  \
        ("A" <= buffer[index] and "Z" >= buffer[index]) or  \
        ("0" <= buffer[index] and "9" >= buffer[index]) or  \
        ("_" == buffer[index])): index = index + 1
    return index, buffer[start: index]

#   解析器
class parser_wrap_t:
    def __init__(self, func = None, name = None):
        self.func = func
        self.name = name
        self.subs = []
    
    def parse(self, buffer, index, length):
        return self.func(buffer, index, length, self.subs)

    def get_type_name(self):
        if   self.func == parse_to_int:
            return "i"
        elif self.func == parse_to_str:
            return "s"
        elif self.func == parse_to_bool:
            return "b"
        elif self.func == parse_to_float:
            return "f"
        elif self.func == parse_to_struct:
            return "t"
        elif self.func == parse_to_list:
            return "list"
        elif self.func == parse_to_dict:
            return "dict"
#
#   解析器
#
def parse_to_int(buffer, index, length, subs):
    result = []
    while index != length:
        if "0" <= buffer[index] and "9" >= buffer[index]:
            result.append(buffer[index]); index += 1
        elif len(result) == 0   and "-" == buffer[index]:
            result.append(buffer[index]); index += 1
        else: break
    return index, "".join(result)

def parse_to_str(buffer, index, length, subs):
    assert buffer[index] == "\"", "str"
    index = index + 1
    match = False
    result = []
    while index != length:
        if buffer[index] == "\\": match = True
        if buffer[index] == "\"" and not match:
            break
        result.append(buffer[index])
        index = index + 1; match = False
    assert buffer[index] == "\"", "str"
    return index + 1, "\"" + "".join(result) + "\""

def parse_to_bool(buffer, index, length, subs):
    assert buffer[index] == "0" or buffer[index] == "1", "bool"
    return index + 1, "false" if buffer[index] == "0" else "true"

def parse_to_float(buffer, index, length, subs):
    result = []
    dot = False
    while index != length:
        if "0" <= buffer[index] and "9" >= buffer[index]:
            result.append(buffer[index]); index = index + 1
        elif len(result) == 0   and "-" == buffer[index]:
            result.append(buffer[index]); index = index + 1
        elif not dot and "." == buffer[index]:
            result.append(buffer[index]); index = index + 1
        else: break
    return index, "".join(result)

def parse_to_list(buffer, index, length, subs):
    assert buffer[index] == "[", "list"
    result = []
    while index != length and buffer[index] != "]":
        index = skip_space_char(buffer, index + 1, length)
        index, item = subs[0].parse(buffer, index, length)
        index = skip_space_char(buffer, index, length)
        result.append(item)
        assert buffer[index] == "," or buffer[index] == "]", "list"

    assert buffer[index] == "]", "list"
    return index + 1, "[" + ", ".join(result) + "]"

def parse_to_dict(buffer, index, length, subs):
    assert buffer[index] == "{", "dict"
    result = []
    while index != length and buffer[index] != "}":
        index = skip_space_char(buffer, index + 1, length)
        index, key = subs[0].parse(buffer, index, length)
        index = skip_space_char(buffer, index, length)
        assert buffer[index] == ":", "dict"
        index = skip_space_char(buffer, index + 1, length)
        index, item = subs[1].parse(buffer, index, length)
        result.append("%s: %s" %(key, item))
        assert buffer[index] == "," or buffer[index] == "}", "dict"

    assert buffer[index] == "}", "dict"
    return index + 1, "{" + ", ".join(result) + "}"

def parse_to_struct(buffer, index, length, subs):
    assert buffer[index] == "<", "struct"
    result = []
    pindex = 0
    while index != length and buffer[index] != ">":
        index += 1
        index = skip_space_char(buffer, index, length)
        index, value = subs[pindex].parse(buffer, index, length)
        index = skip_space_char(buffer, index, length)
        result.append("\"%s\": %s" % (subs[pindex].name, value))
        pindex = pindex + 1
        assert buffer[index] == "," or buffer[index] == ">", "struct"
    
    assert buffer[index] == ">", "struct"
    return index + 1, "{" + ", ".join(result) + "}"

def gen_parser_wrap(buffer, index, length):
    parser_wrap = None
    if match_prefix(buffer, "i", index):
        index += len("i")
        parser_wrap = parser_wrap_t(parse_to_int)
    elif match_prefix(buffer, "s", index):
        index += len("s")
        parser_wrap = parser_wrap_t(parse_to_str)
    elif match_prefix(buffer, "b", index):
        index += len("b")
        parser_wrap = parser_wrap_t(parse_to_bool)
    elif match_prefix(buffer, "f", index):
        index += len("f")
        parser_wrap = parser_wrap_t(parse_to_float)
    elif buffer[index] == "[":
        index = index + 1
        index = skip_space_char(buffer, index, length)
        index, val_parser = gen_parser_wrap(buffer, index, length)
        index = skip_space_char(buffer, index, length)
        parser_wrap = parser_wrap_t(parse_to_list)
        parser_wrap.subs.append(val_parser)
        assert buffer[index] == "]", "list"
        index = index + 1
    elif buffer[index] == "{":
        index = index + 1
        key_parser = parser_wrap_t(parse_to_str)
        index = skip_space_char(buffer, index, length)
        index, val_parser = gen_parser_wrap(buffer, index, length)
        index = skip_space_char(buffer, index, length)
        parser_wrap = parser_wrap_t(parse_to_dict)
        parser_wrap.subs.append(key_parser)
        parser_wrap.subs.append(val_parser)
        assert buffer[index] == "}", "dict"
        index = index + 1
    elif buffer[index] == "<":
        parser_wrap = parser_wrap_t(parse_to_struct)
        while index != length:
            if buffer[index] == ">": break
            index = index + 1
            index = skip_space_char(buffer, index, length)
            index, val_parser = gen_parser_wrap(buffer, index, length)
            index = skip_space_char(buffer, index, length)
            parser_wrap.subs.append(val_parser)
            assert buffer[index] == "," or buffer[index] == ">", "struct"
        assert buffer[index] == ">", "struct"
        index = index + 1

    index  = skip_space_char(buffer, index, length)
    index, k = get_type_name(buffer, index, length)
    if parser_wrap != None: parser_wrap.name = k
    return index, parser_wrap

#   初始化解析器
def get_parser_wraps(xl, row, output):
    for col in range(1, xl.max_column + 1):
        try:
            buffer = xl_value(xl, row, col)
            output.append(gen_parser_wrap(buffer, 0, len(buffer))[1])
        except:
            assert False, "%d | %d | %s" % (row, col, buffer)
    return row + 1

#   解析单元格
def get_output_lines(xlsx, row, parser_wrap_list, output_line_list):
    collect_keys = []
    for row in range(row, xlsx.max_row + 1):
        if xl_value(xlsx, row, 1) == "//":
            continue

        lines = []
        for col in range(1, xlsx.max_column + 1):
            try:
                value  = xl_value(xlsx, row, col)
                parser = parser_wrap_list[col -1]
                if parser == None: continue
                _, v = parser.parse(value, 0, len(value))
                lines.append("\"%s\": %s" % (parser.name, v))
            except AssertionError as e:
                assert False, "%d | %d | %s | %s" % (row, col, e, value)
        key = xl_value(xlsx, row, 1)
        val = "{" + ", ".join(lines) + "}"
        output_line_list.append("%s: %s" % (to_key(key), val))
        collect_keys.append(key)

    fmt = "\"__Index__\": [ %s ]"
    str = ", ".join(collect_keys)
    output_line_list.append(fmt % str)
    return row

#   导出Json
def to_json(file_name):
    #   读第一张表
    xlsx = openpyxl.load_workbook(file_name, data_only = True)
    name, xlsx = xlsx.sheetnames[0], xlsx[xlsx.sheetnames[0]]

    try:
        row = skip_comment_line(xlsx, 1)
        parser_cols = []
        output_rows = []
        row = get_parser_wraps(xlsx, row, parser_cols)
        row = get_output_lines(xlsx, row, parser_cols, output_rows)
        return (name, "{\n" + ",\n".join(output_rows) + "\n}", parser_cols)
    except AssertionError as e:
        assert False, "%s | %s" % (file_name, e)
